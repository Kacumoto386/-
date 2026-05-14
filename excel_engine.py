"""
Excel读写引擎 v2.0
封装openpyxl操作，支持12个Sheet页的高级读写、格式化、公式写入等功能
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from datetime import datetime, date, timedelta
import os
import re
from copy import copy


class ExcelEngine:
    """Excel文件的底层读写操作"""

    def __init__(self, filepath):
        self.filepath = filepath
        self.wb = None
        self.load()

    # ==================== 基础操作 ====================

    def load(self):
        """加载工作簿"""
        if os.path.exists(self.filepath):
            self.wb = openpyxl.load_workbook(self.filepath)
        else:
            self.wb = openpyxl.Workbook()
            # 删除默认Sheet
            if "Sheet" in self.wb.sheetnames:
                del self.wb["Sheet"]

    def save(self):
        """保存工作簿"""
        self.wb.save(self.filepath)

    def close(self):
        """关闭工作簿"""
        if self.wb:
            self.wb.close()

    def get_sheet(self, sheet_name):
        """获取Sheet页，不存在则创建"""
        if sheet_name in self.wb.sheetnames:
            return self.wb[sheet_name]
        return self.wb.create_sheet(title=sheet_name)

    def sheet_exists(self, name):
        """检查Sheet是否存在"""
        return name in self.wb.sheetnames

    def rename_sheet(self, old_name, new_name):
        """重命名Sheet"""
        if old_name in self.wb.sheetnames:
            ws = self.wb[old_name]
            ws.title = new_name
            return True
        return False

    def delete_sheet(self, name):
        """删除Sheet"""
        if name in self.wb.sheetnames:
            del self.wb[name]
            return True
        return False

    # ==================== 单元格读写 ====================

    def read_cell(self, sheet_name, row, col):
        """读取单元格"""
        ws = self.get_sheet(sheet_name)
        return ws.cell(row=row, column=col).value

    def write_cell(self, sheet_name, row, col, value):
        """写入单元格"""
        ws = self.get_sheet(sheet_name)
        ws.cell(row=row, column=col).value = value
        return True

    def read_range(self, sheet_name, min_row, max_row, min_col, max_col):
        """读取区域数据"""
        ws = self.get_sheet(sheet_name)
        data = []
        for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                                min_col=min_col, max_col=max_col, values_only=True):
            data.append(list(row))
        return data

    def write_range(self, sheet_name, start_row, start_col, data):
        """写入区域数据"""
        ws = self.get_sheet(sheet_name)
        for i, row_data in enumerate(data):
            for j, value in enumerate(row_data):
                ws.cell(row=start_row + i, column=start_col + j).value = value
        self.save()
        return True

    # ==================== 行操作 ====================

    def get_data_start_row(self):
        """获取数据起始行（前3行为标题/表头）"""
        return 4

    def find_next_empty_row(self, sheet_name, col=1):
        """找到指定列的下一个空行"""
        ws = self.get_sheet(sheet_name)
        row = self.get_data_start_row()
        while ws.cell(row=row, column=col).value is not None:
            row += 1
        return row

    def get_last_data_row(self, sheet_name, col=1):
        """获取最后一行的行号"""
        ws = self.get_sheet(sheet_name)
        row = ws.max_row
        start = self.get_data_start_row()
        while row >= start:
            if ws.cell(row=row, column=col).value is not None:
                return row
            row -= 1
        return start - 1

    def get_all_data(self, sheet_name, key_col=1):
        """获取所有数据行（返回列表，每行为字典）"""
        headers = self.get_headers(sheet_name)
        ws = self.get_sheet(sheet_name)
        results = []
        start = self.get_data_start_row()
        end = self.get_last_data_row(sheet_name, key_col)

        for row_num in range(start, end + 1):
            row_data = {}
            has_value = False
            for i, header in enumerate(headers):
                col = i + 1
                val = ws.cell(row=row_num, column=col).value
                row_data[header] = val
                if val is not None and str(val).strip():
                    has_value = True
            if has_value:
                row_data["_row"] = row_num
                results.append(row_data)
        return results

    def write_row(self, sheet_name, row_num, data_dict):
        """将字典写入一行（以headers为列映射）"""
        headers = self.get_headers(sheet_name)
        ws = self.get_sheet(sheet_name)
        for i, header in enumerate(headers):
            col = i + 1
            if header in data_dict:
                ws.cell(row=row_num, column=col).value = data_dict[header]
        self.save()
        return True

    def update_row(self, sheet_name, row_num, data_dict):
        """更新一行数据"""
        return self.write_row(sheet_name, row_num, data_dict)

    def delete_row(self, sheet_name, row_num):
        """删除一行（物理删除）"""
        ws = self.get_sheet(sheet_name)
        ws.delete_rows(row_num)
        self.save()
        return True

    def append_row(self, sheet_name, data_dict):
        """追加一行数据"""
        row = self.find_next_empty_row(sheet_name)
        return self.write_row(sheet_name, row, data_dict)

    # ==================== 查询操作 ====================

    def find_rows(self, sheet_name, conditions):
        """
        按条件查找行
        conditions: {"字段名": 值} 或 {"字段名": lambda x: x > 5}
        返回匹配的行列表
        """
        data = self.get_all_data(sheet_name)
        results = []
        for row in data:
            match = True
            for key, cond in conditions.items():
                val = row.get(key)
                if callable(cond):
                    if not cond(val):
                        match = False
                        break
                else:
                    if val != cond:
                        match = False
                        break
            if match:
                results.append(row)
        return results

    def row_to_dict(self, sheet_name, row_num):
        """将一行转为字典"""
        headers = self.get_headers(sheet_name)
        ws = self.get_sheet(sheet_name)
        row_data = {"_row": row_num}
        for i, header in enumerate(headers):
            col = i + 1
            val = ws.cell(row=row_num, column=col).value
            row_data[header] = val
        return row_data

    def clear_sheet(self, sheet_name, keep_headers=True, headers=None):
        """清空Sheet页数据"""
        ws = self.get_sheet(sheet_name)
        if keep_headers:
            # 保留前3行（标题行+表头行），清除后续数据
            if headers:
                # 用新表头重新初始化
                for i, h in enumerate(headers, 1):
                    ws.cell(row=3, column=i).value = h
            # 清除数据行
            start = self.get_data_start_row()
            max_row = ws.max_row
            max_col = ws.max_column
            for row in range(start, max_row + 1):
                for col in range(1, max_col + 1):
                    ws.cell(row=row, column=col).value = None
        else:
            # 完全清空
            for row in ws.iter_rows():
                for cell in row:
                    cell.value = None
        self.save()
        return True

    def format_header_row(self, sheet_name, row, max_col):
        """格式化表头行（别名方法）"""
        self.apply_header_style(sheet_name, row, max_col)

    def search_rows(self, sheet_name, keyword, fields=None):
        """
        模糊搜索
        keyword: 搜索关键词
        fields: 搜索字段列表，None表示所有字段
        """
        data = self.get_all_data(sheet_name)
        results = []
        keyword = str(keyword).lower()
        for row in data:
            for key, val in row.items():
                if key == "_row":
                    continue
                if fields and key not in fields:
                    continue
                if val is not None and keyword in str(val).lower():
                    results.append(row)
                    break
        return results

    # ==================== 表头操作 ====================

    def get_headers(self, sheet_name, row=3):
        """获取表头"""
        ws = self.get_sheet(sheet_name)
        headers = []
        col = 1
        while True:
            val = ws.cell(row=row, column=col).value
            if val is None:
                break
            headers.append(str(val).strip())
            col += 1
        return headers

    def get_header_col(self, sheet_name, header_name):
        """获取表头对应的列号"""
        headers = self.get_headers(sheet_name)
        for i, h in enumerate(headers):
            if h == header_name:
                return i + 1
        return None

    # ==================== 样式操作 ====================

    def apply_header_style(self, sheet_name, row, max_col):
        """应用表头样式"""
        ws = self.get_sheet(sheet_name)
        header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def apply_data_style(self, sheet_name, row, max_col):
        """应用数据行样式"""
        ws = self.get_sheet(sheet_name)
        data_font = Font(name="微软雅黑", size=10)
        data_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

    def set_column_widths(self, sheet_name, widths_dict):
        """设置列宽"""
        ws = self.get_sheet(sheet_name)
        for col_letter, width in widths_dict.items():
            ws.column_dimensions[col_letter].width = width

    def set_cell_style(self, sheet_name, row, col, font=None, fill=None,
                       alignment=None, border=None, number_format=None):
        """设置单元格样式"""
        ws = self.get_sheet(sheet_name)
        cell = ws.cell(row=row, column=col)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            cell.border = border
        if number_format:
            cell.number_format = number_format

    def merge_cells(self, sheet_name, start_row, start_col, end_row, end_col):
        """合并单元格"""
        ws = self.get_sheet(sheet_name)
        ws.merge_cells(
            start_row=start_row, start_column=start_col,
            end_row=end_row, end_column=end_col
        )

    def apply_conditional_format_alert(self, sheet_name, cell_range, threshold,
                                       fill_color="FF0000", font_color="FFFFFF"):
        """应用条件格式：值 <= 阈值 时标红"""
        ws = self.get_sheet(sheet_name)
        red_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        red_font = Font(color=font_color, bold=True)
        ws.conditional_formatting.add(
            cell_range,
            CellIsRule(operator='lessThanOrEqual', formula=[str(threshold)],
                       fill=red_fill, font=red_font)
        )

    # ==================== 数据验证（下拉） ====================

    def add_dropdown(self, sheet_name, cell_range, options):
        """添加下拉列表数据验证"""
        ws = self.get_sheet(sheet_name)
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(options)}"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="输入错误",
            error="请从下拉列表中选择有效值",
        )
        ws.add_data_validation(dv)
        dv.add(cell_range)
        return dv

    def add_dropdown_from_sheet(self, sheet_name, cell_range,
                                source_sheet, source_col, source_start=4):
        """添加从另一个Sheet读取的下拉列表"""
        ws = self.get_sheet(sheet_name)
        source_ref = f"'{source_sheet}'!${source_col}${source_start}:${source_col}$1000"
        dv = DataValidation(
            type="list",
            formula1=source_ref,
            allow_blank=True,
        )
        ws.add_data_validation(dv)
        dv.add(cell_range)
        return dv

    # ==================== 公式辅助 ====================

    @staticmethod
    def formula_vlookup(lookup_value, table_range, col_index, exact=True):
        """生成VLOOKUP公式"""
        return f'=VLOOKUP({lookup_value},{table_range},{col_index},{"FALSE" if exact else "TRUE"})'

    @staticmethod
    def formula_sumifs(sum_range, criteria_range1, criteria1, criteria_range2=None, criteria2=None):
        """生成SUMIFS公式"""
        if criteria_range2:
            return f'=SUMIFS({sum_range},{criteria_range1},{criteria1},{criteria_range2},{criteria2})'
        return f'=SUMIF({sum_range},{criteria_range1},{criteria1})'

    @staticmethod
    def formula_datedif(start_date, end_date, unit="Y"):
        """生成DATEDIF公式"""
        return f'=DATEDIF({start_date},{end_date},"{unit}")'

    @staticmethod
    def formula_if(condition, true_val, false_val):
        """生成IF公式"""
        return f'=IF({condition},{true_val},{false_val})'

    @staticmethod
    def formula_today():
        """生成TODAY公式"""
        return "=TODAY()"

    @staticmethod
    def formula_now():
        """生成NOW公式"""
        return "=NOW()"

    # ==================== 高级功能 ====================

    def sheet_to_dicts(self, sheet_name, key_field=None):
        """
        将Sheet转为字典列表（key_field指定则返回{key: row_dict}）
        """
        rows = self.get_all_data(sheet_name)
        if key_field and rows:
            return {row.get(key_field): row for row in rows if row.get(key_field)}
        return rows

    def get_summary_stats(self, sheet_name, key_col=1, value_col=None):
        """
        简单汇总统计
        返回: {"total_rows": N, "total_value": sum}
        """
        data = self.get_all_data(sheet_name, key_col)
        total = len(data)
        total_value = 0
        if value_col:
            headers = self.get_headers(sheet_name)
            val_idx = headers.index(value_col) if value_col in headers else -1
            if val_idx >= 0:
                for row in data:
                    val = row.get(value_col, 0)
                    try:
                        total_value += float(val or 0)
                    except (ValueError, TypeError):
                        pass
        return {"total_rows": total, "total_value": total_value}

    def backup(self):
        """创建当前工作簿的备份"""
        if not os.path.exists(self.filepath):
            return None
        backup_dir = os.path.join(os.path.dirname(self.filepath), "backup")
        os.makedirs(backup_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        basename = os.path.basename(self.filepath)
        backup_path = os.path.join(backup_dir, f"{basename}.{timestamp}.bak")
        self.save()
        import shutil
        shutil.copy2(self.filepath, backup_path)
        self.load()
        return backup_path
