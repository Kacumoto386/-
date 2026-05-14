"""
初始化工作簿 - 创建完整的健身房管理系统Excel文件
"""
import os
import sys
from datetime import datetime, date, timedelta

# 确保项目根目录在Path中
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config import (
    VERSION_TAG,
    SHEETS, HEADERS, DATA_DIR, EXCEL_PATH, DATA_START_ROW,
    MEMBER_LEVELS, GENDERS, MEMBER_STATUSES, MEMBER_EXTRA_COLS,
    STAFF_POSITIONS,
    COURSE_TYPES, COURSE_STATUSES,
    PAYMENT_METHODS,
    CLASS_STATUSES,
    RECHARGE_TYPES,
    BOOKING_STATUSES, BOOKING_TIME_SLOTS,
    PACKAGE_STATUSES,
    OP_TYPES, OP_MODULES,
)
from core.excel_engine import ExcelEngine


class WorkbookInitializer:
    """初始化Excel工作簿"""

    def __init__(self, engine):
        self.engine = engine
        self.wb = engine.wb

    def initialize(self, with_sample_data=True):
        """完整初始化工作簿"""
        print("🚀 开始初始化工作簿...")

        # 删除默认Sheet
        for name in list(self.wb.sheetnames):
            del self.wb[name]

        # 创建所有Sheet页
        for key in SHEETS:
            sheet_name = SHEETS[key]
            self.wb.create_sheet(title=sheet_name)
            print(f"  ✅ 创建Sheet: {sheet_name}")

        # 初始化各Sheet
        """完整初始化工作簿"""
        print("🚀 开始初始化工作簿...")

        # 删除默认Sheet
        for name in list(self.wb.sheetnames):
            del self.wb[name]

        # 创建12个Sheet页
        for key in SHEETS:
            sheet_name = SHEETS[key]
            self.wb.create_sheet(title=sheet_name)
            print(f"  ✅ 创建Sheet: {sheet_name}")

        # 初始化各Sheet
        self._init_dashboard()
        self._init_member()
        self._init_staff()
        self._init_course()
        self._init_sale()
        self._init_class_record()
        self._init_recharge()
        self._init_booking()
        self._init_lesson_package()
        self._init_body_measurement()
        self._init_product()
        self._init_product_sale()
        self._init_stat_sale()
        self._init_stat_class()
        self._init_stat_commission()
        self._init_contract()
        self._init_alert()
        self._init_log()

        # 添加示例数据
        if with_sample_data:
            self._add_sample_data()

        # 保存
        self.engine.save()
        print(f"\n✅ 工作簿初始化完成！")
        print(f"📁 文件路径: {self.engine.filepath}")
        print(f"📊 Sheet页数: {len(self.wb.sheetnames)}")
        print(f"📋 名称: {[s.title for s in self.wb.worksheets]}")

    # ==================== Sheet初始化 ====================

    def _init_dashboard(self):
        """首页看板"""
        ws = self.engine.get_sheet(SHEETS["dashboard"])
        title_row, sub_row, header_row = 1, 2, 3

        # 标题
        ws.cell(row=title_row, column=1, value=VERSION_TAG)
        ws.merge_cells(start_row=title_row, start_column=1,
                       end_row=title_row, end_column=8)
        ws.cell(row=title_row, column=1).font = self._title_font()

        # 核心指标标题
        indicators = [
            "核心指标", "", "", "", "", "快速入口", "", ""
        ]
        for i, v in enumerate(indicators, 1):
            ws.cell(row=sub_row, column=i, value=v)
        ws.cell(row=sub_row, column=1).font = self._subtitle_font()
        ws.cell(row=sub_row, column=6).font = self._subtitle_font()

        # 指标项
        metrics = [
            ("当前有效会员数", ""), ("本月新增会员", ""),
            ("当前总剩余课时", ""), ("本月上课总节数", ""),
            ("本月售课总金额", ""), ("本月到期会员数", ""),
        ]
        for i, (label, value) in enumerate(metrics):
            row = header_row + i
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=1).font = self._label_font()
            ws.cell(row=row, column=2, value=value)

        # 快速入口
        quick_links = ["新增会员", "新增售课", "新增上课记录", "到期提醒", "统计报表", "操作日志"]
        for i, link in enumerate(quick_links):
            row = header_row + i
            ws.cell(row=row, column=6, value=link)
            ws.cell(row=row, column=6).font = self._link_font()

        # 图表区标题
        chart_row = header_row + 8
        ws.cell(row=chart_row, column=1, value="📊 数据可视化区")
        ws.cell(row=chart_row, column=1).font = self._subtitle_font()
        ws.merge_cells(start_row=chart_row, start_column=1,
                       end_row=chart_row, end_column=10)

        # 图表占位
        chart_labels = [
            ("近6个月售课金额趋势图", 4, 1), ("各课程上课占比", 4, 6),
            ("员工售课排行榜Top5", 10, 1),
        ]
        for label, c_row, c_col in chart_labels:
            ws.cell(row=chart_row + c_row, column=c_col, value=label)
            ws.cell(row=chart_row + c_row, column=c_col).font = self._label_font()
            # 合并区域模拟图表占位
            ws.merge_cells(start_row=chart_row + c_row, start_column=c_col,
                           end_row=chart_row + c_row + 3, end_column=c_col + 4)

        # 设置列宽
        for col, width in {1: 20, 2: 15, 3: 15, 4: 15, 5: 15,
                           6: 20, 7: 15, 8: 15, 9: 15, 10: 15}.items():
            ws.column_dimensions[chr(64 + col)].width = width

    def _init_member(self):
        """会员信息"""
        ws = self.engine.get_sheet(SHEETS["member"])
        self._write_title(ws, "会员信息管理")
        self._write_headers(ws, "member")
        self._add_dropdowns(ws, "member", {
            3: GENDERS,     # 性别
            8: MEMBER_LEVELS,       # 会员等级
            11: MEMBER_STATUSES,      # 会员状态
            24: MEMBER_STATUSES,      # 跟进状态
        })
        self.engine.apply_header_style(SHEETS["member"], DATA_START_ROW - 1,
                                       len(HEADERS["member"]))
        self._set_col_widths(ws, "member",
                             {1: 16, 2: 10, 3: 6, 4: 15, 5: 12, 6: 6,
                              7: 20, 8: 8, 9: 12, 10: 12, 11: 10,
                              12: 12, 13: 10, 14: 10, 15: 10,
                              16: 10, 17: 10, 18: 14, 19: 14,
                              20: 10, 21: 12, 22: 10, 23: 20, 24: 10,
                              25: 14, 26: 20})

    def _init_staff(self):
        """员工信息"""
        ws = self.engine.get_sheet(SHEETS["staff"])
        self._write_title(ws, "员工信息管理")
        self._write_headers(ws, "staff")
        self._add_dropdowns(ws, "staff", {
            3: GENDERS,       # 性别
            4: STAFF_POSITIONS,     # 岗位
            10: MEMBER_STATUSES,       # 员工状态
        })
        self.engine.apply_header_style(SHEETS["staff"], DATA_START_ROW - 1,
                                       len(HEADERS["staff"]))
        self._set_col_widths(ws, "staff",
                             {1: 14, 2: 10, 3: 6, 4: 12, 5: 12, 6: 12,
                              7: 6, 8: 15, 9: 20, 10: 8, 11: 12,
                              12: 20, 13: 20,
                              14: 12, 15: 12, 16: 10,
                              17: 14, 18: 14, 19: 14, 20: 14,
                              21: 14, 22: 14, 23: 14, 24: 12})

    def _init_course(self):
        """课程&运动项目"""
        ws = self.engine.get_sheet(SHEETS["course"])
        self._write_title(ws, "课程 & 运动项目管理")
        self._write_headers(ws, "course")
        self._add_dropdowns(ws, "course", {
            3: COURSE_TYPES,              # 运动项目
            4: COURSE_TYPES,        # 课程类型
            11: MEMBER_LEVELS,      # 适用会员等级（简化为单选项）
            12: MEMBER_STATUSES,             # 是否支持试课
            15: COURSE_STATUSES,      # 课程状态
        })
        self.engine.apply_header_style(SHEETS["course"], DATA_START_ROW - 1,
                                       len(HEADERS["course"]))
        self._set_col_widths(ws, "course",
                             {1: 14, 2: 20, 3: 14, 4: 16,
                              5: 16, 6: 12, 7: 12, 8: 12,
                              9: 14, 10: 14, 11: 14,
                              12: 12, 13: 12, 14: 30, 15: 10})

    def _init_sale(self):
        """售课记录"""
        ws = self.engine.get_sheet(SHEETS["sale"])
        self._write_title(ws, "售课记录管理")
        self._write_headers(ws, "sale")
        self._add_dropdowns(ws, "sale", {
            3: STAFF_POSITIONS,     # 销售员工（简化）
            18: PAYMENT_METHODS,    # 付款方式
            24: MEMBER_STATUSES,             # 是否已开票
            27: MEMBER_STATUSES,       # 购课来源
        })
        self.engine.apply_header_style(SHEETS["sale"], DATA_START_ROW - 1,
                                       len(HEADERS["sale"]))
        self._set_col_widths(ws, "sale",
                             {1: 18, 2: 12, 3: 10, 4: 16, 5: 10,
                              6: 12, 7: 14, 8: 20, 9: 14, 10: 16,
                              11: 10, 12: 10, 13: 10, 14: 10, 15: 8,
                              16: 12, 17: 12, 18: 12,
                              19: 12, 20: 14, 21: 14, 22: 14,
                              23: 14, 24: 10, 25: 12, 26: 16,
                              27: 12, 28: 14, 29: 20})

    def _init_class_record(self):
        """会员上课记录"""
        ws = self.engine.get_sheet(SHEETS["class_record"])
        self._write_title(ws, "会员上课记录管理")
        self._write_headers(ws, "class_record")
        self._add_dropdowns(ws, "class_record", {
            14: CLASS_STATUSES,       # 上课状态
            17: ["非常满意", "满意", "一般", "不满意", "未评价"],       # 上课评价
        })
        self.engine.apply_header_style(SHEETS["class_record"], DATA_START_ROW - 1,
                                       len(HEADERS["class_record"]))
        self._set_col_widths(ws, "class_record",
                             {1: 18, 2: 12, 3: 10, 4: 10,
                              5: 10, 6: 10, 7: 16, 8: 10,
                              9: 14, 10: 20, 11: 14, 12: 16,
                              13: 10, 14: 10, 15: 12, 16: 12,
                              17: 10, 18: 20, 19: 20, 20: 30,
                              21: 20, 22: 14, 23: 14,
                              24: 12, 25: 16, 26: 20})

    def _init_recharge(self):
        """会员充值记录"""
        ws = self.engine.get_sheet(SHEETS["recharge"])
        self._write_title(ws, "会员充值记录管理")
        self._write_headers(ws, "recharge")
        self._add_dropdowns(ws, "recharge", {
            8: PAYMENT_METHODS,     # 付款方式
            9: RECHARGE_TYPES,      # 充值类型
        })
        self.engine.apply_header_style(SHEETS["recharge"], DATA_START_ROW - 1,
                                       len(HEADERS["recharge"]))
        self._set_col_widths(ws, "recharge",
                             {1: 18, 2: 12, 3: 16, 4: 10,
                              5: 12, 6: 12, 7: 12, 8: 12,
                              9: 12, 10: 12, 11: 10, 12: 20})


    def _init_booking(self):
        """初始化预约管理"""
        ws = self.engine.get_sheet(SHEETS["booking"])
        self._write_title(ws, "预约管理")
        self._write_headers(ws, "booking")
        self._set_col_widths(ws, "booking",
                       {1: 14, 2: 12, 3: 10, 4: 10, 5: 14, 6: 10, 7: 14,
                        8: 14, 9: 16, 10: 14, 11: 10, 12: 14, 13: 10, 14: 10})

    def _init_lesson_package(self):
        """初始化课程包管理"""
        ws = self.engine.get_sheet(SHEETS["lesson_package"])
        self._write_title(ws, "课程包管理")
        self._write_headers(ws, "lesson_package")
        self._set_col_widths(ws, "lesson_package",
                       {1: 14, 2: 14, 3: 14, 4: 10, 5: 14, 6: 16,
                        7: 10, 8: 12, 9: 10, 10: 12, 11: 12, 12: 10})

    def _init_body_measurement(self):
        """初始化体测记录"""
        ws = self.engine.get_sheet(SHEETS["body_measurement"])
        self._write_title(ws, "体测记录")
        self._write_headers(ws, "body_measurement")
        self._set_col_widths(ws, "body_measurement",
                       {1: 14, 2: 14, 3: 10, 4: 12, 5: 10, 6: 10,
                        7: 10, 8: 10, 9: 12, 10: 14, 11: 10, 12: 20})

    def _init_product(self):
        """初始化商品管理"""
        ws = self.engine.get_sheet(SHEETS["product"])
        self._write_title(ws, "商品管理")
        self._write_headers(ws, "product")
        self._set_col_widths(ws, "product",
                       {1: 14, 2: 20, 3: 12, 4: 10, 5: 10, 6: 10,
                        7: 6, 8: 16, 9: 20})

    def _init_product_sale(self):
        """初始化商品零售"""
        ws = self.engine.get_sheet(SHEETS["product_sale"])
        self._write_title(ws, "商品零售")
        self._write_headers(ws, "product_sale")
        self._set_col_widths(ws, "product_sale",
                       {1: 14, 2: 12, 3: 14, 4: 10, 5: 20, 6: 8,
                        7: 8, 8: 10, 9: 12, 10: 10, 11: 20})

    def _init_contract(self):
        """初始化合同管理"""
        ws = self.engine.get_sheet(SHEETS["contract"])
        self._write_title(ws, "合同管理")
        self._write_headers(ws, "contract")
        self._set_col_widths(ws, "contract",
                       {1: 14, 2: 14, 3: 14, 4: 10, 5: 12,
                        6: 30, 7: 12, 8: 12, 9: 12, 10: 10,
                        11: 10, 12: 14, 13: 10, 14: 20, 15: 20})

    def _init_stat_sale(self):
        """售课统计"""
        ws = self.engine.get_sheet(SHEETS["stat_sale"])
        self._write_title(ws, "📊 售课统计报表")

        # 按课程统计
        ws.cell(row=2, column=1, value="一、按课程统计")
        ws.cell(row=2, column=1).font = self._subtitle_font()
        stat_sale_headers = [
            "课程编号", "课程名称", "本月售课笔数", "本月售课金额",
            "本月售课课时", "累计售课笔数", "累计售课金额", "累计售课课时",
        ]
        for i, h in enumerate(stat_sale_headers, 1):
            ws.cell(row=3, column=i, value=h)
        self.engine.apply_header_style(SHEETS["stat_sale"], 3, len(stat_sale_headers))

        # 按销售员工统计
        staff_row = 13
        ws.cell(row=staff_row, column=1, value="二、按销售员工统计")
        ws.cell(row=staff_row, column=1).font = self._subtitle_font()
        staff_headers = [
            "员工编号", "姓名", "本月售课笔数", "本月售课金额",
            "本月提成", "累计售课笔数", "累计售课金额", "累计提成",
        ]
        for i, h in enumerate(staff_headers, 1):
            ws.cell(row=staff_row + 1, column=i, value=h)
        self.engine.apply_header_style(SHEETS["stat_sale"], staff_row + 1, len(staff_headers))

        # 按月份统计
        month_row = 23
        ws.cell(row=month_row, column=1, value="三、按月份统计")
        ws.cell(row=month_row, column=1).font = self._subtitle_font()
        month_headers = [
            "月份", "售课金额", "售课课时", "新增会员", "续费会员数", "续费率",
        ]
        for i, h in enumerate(month_headers, 1):
            ws.cell(row=month_row + 1, column=i, value=h)
        self.engine.apply_header_style(SHEETS["stat_sale"], month_row + 1, len(month_headers))

        self._set_col_widths(ws, "stat_sale",
                             {1: 14, 2: 20, 3: 14, 4: 14,
                              5: 14, 6: 14, 7: 14, 8: 14})

    def _init_stat_class(self):
        """上课统计"""
        ws = self.engine.get_sheet(SHEETS["stat_class"])
        self._write_title(ws, "📊 上课统计报表")

        # 按课程统计
        ws.cell(row=2, column=1, value="一、按课程统计")
        ws.cell(row=2, column=1).font = self._subtitle_font()
        headers1 = [
            "课程编号", "课程名称", "运动项目", "本月上课节数",
            "本月消耗课时", "本月上课人次", "本月评价均分",
            "累计上课节数", "累计消耗课时",
        ]
        for i, h in enumerate(headers1, 1):
            ws.cell(row=3, column=i, value=h)
        self.engine.apply_header_style(SHEETS["stat_class"], 3, len(headers1))

        # 按教练统计
        coach_row = 13
        ws.cell(row=coach_row, column=1, value="二、按教练统计")
        ws.cell(row=coach_row, column=1).font = self._subtitle_font()
        headers2 = [
            "教练编号", "姓名", "本月上课节数", "本月提成",
            "本月服务会员数", "本月评价均分", "累计上课节数", "累计提成",
        ]
        for i, h in enumerate(headers2, 1):
            ws.cell(row=coach_row + 1, column=i, value=h)
        self.engine.apply_header_style(SHEETS["stat_class"], coach_row + 1, len(headers2))

        # 按会员统计
        member_row = 23
        ws.cell(row=member_row, column=1, value="三、按会员统计")
        ws.cell(row=member_row, column=1).font = self._subtitle_font()
        headers3 = [
            "会员编号", "姓名", "本月上课次数", "本月消耗课时",
            "常练课程", "最近上课日期", "本月评价均分",
        ]
        for i, h in enumerate(headers3, 1):
            ws.cell(row=member_row + 1, column=i, value=h)
        self.engine.apply_header_style(SHEETS["stat_class"], member_row + 1, len(headers3))

        self._set_col_widths(ws, "stat_class",
                             {1: 14, 2: 12, 3: 14, 4: 14,
                              5: 14, 6: 14, 7: 14, 8: 14, 9: 14})

    def _init_stat_commission(self):
        """员工提成汇总"""
        ws = self.engine.get_sheet(SHEETS["stat_commission"])
        self._write_title(ws, "📊 员工提成汇总报表")

        ws.cell(row=2, column=1, value="一、本月员工提成汇总")
        ws.cell(row=2, column=1).font = self._subtitle_font()
        headers = [
            "员工编号", "姓名", "岗位", "本月售课提成", "本月上课提成",
            "本月总提成", "本月底薪", "本月应发合计",
            "累计售课提成", "累计上课提成", "累计总提成",
        ]
        for i, h in enumerate(headers, 1):
            ws.cell(row=3, column=i, value=h)
        self.engine.apply_header_style(SHEETS["stat_commission"], 3, len(headers))

        # 按月份统计（每个员工）
        month_row = 13
        ws.cell(row=month_row, column=1, value="二、按月份统计（每人每月）")
        ws.cell(row=month_row, column=1).font = self._subtitle_font()
        row_headers = [
            "月份", "员工编号", "姓名", "底薪", "售课金额",
            "售课提成", "上课节数", "上课提成", "总提成", "应发合计",
        ]
        for i, h in enumerate(row_headers, 1):
            ws.cell(row=month_row + 1, column=i, value=h)
        self.engine.apply_header_style(SHEETS["stat_commission"], month_row + 1, len(row_headers))

        self._set_col_widths(ws, "stat_commission",
                             {1: 14, 2: 14, 3: 12, 4: 14, 5: 14,
                              6: 14, 7: 12, 8: 14, 9: 14, 10: 16, 11: 14})

    def _init_alert(self):
        """到期提醒"""
        ws = self.engine.get_sheet(SHEETS["alert"])
        self._write_title(ws, "⏰ 到期提醒管理")
        self._write_headers(ws, "alert")
        self._add_dropdowns(ws, "alert", {
            1: OP_TYPES,         # 提醒类型
            9: OP_MODULES,       # 紧急程度
            10: MEMBER_STATUSES,             # 是否已通知
            12: MEMBER_STATUSES,     # 通知方式
        })
        self.engine.apply_header_style(SHEETS["alert"], DATA_START_ROW - 1,
                                       len(HEADERS["alert"]))
        self._set_col_widths(ws, "alert",
                             {1: 16, 2: 14, 3: 16, 4: 10, 5: 15,
                              6: 30, 7: 14, 8: 10, 9: 16,
                              10: 10, 11: 12, 12: 10, 13: 10, 14: 12})

    def _init_log(self):
        """操作日志"""
        ws = self.engine.get_sheet(SHEETS["log"])
        self._write_title(ws, "📝 操作日志")
        self._write_headers(ws, "log")
        self._add_dropdowns(ws, "log", {
            4: OP_TYPES,            # 操作类型
            5: OP_MODULES,          # 操作模块
        })
        self.engine.apply_header_style(SHEETS["log"], DATA_START_ROW - 1,
                                       len(HEADERS["log"]))
        self._set_col_widths(ws, "log",
                             {1: 18, 2: 20, 3: 10, 4: 10,
                              5: 10, 6: 40, 7: 30, 8: 30})

    # ==================== 示例数据 ====================

    def _add_sample_data(self):
        """添加示例数据"""
        self._add_sample_courses()
        self._add_sample_staff()
        self._add_sample_members()
        self._add_sample_sales()
        self._add_sample_class_records()
        print("  ✅ 示例数据已添加")

    def _add_sample_courses(self):
        """添加示例课程"""
        ws = self.engine.get_sheet(SHEETS["course"])
        courses = [
            ["C001", "1对1私教课", "力量训练", "1对1私教", 60, 1, 300, 200, 180, 1, "全部", "否", 0, "专业私教一对一指导", "上架"],
            ["C002", "瑜伽小班课", "瑜伽", "小班课(3-8人)", 60, 0.5, 150, 100, 90, 8, "全部", "是", 0.5, "专业瑜伽指导", "上架"],
            ["C003", "普拉提器械课", "普拉提", "1对1私教", 60, 1, 350, 250, 180, 1, "金卡,钻石", "否", 0, "普拉提核心床训练", "上架"],
            ["C004", "游泳私教课", "游泳", "1对1私教", 60, 1, 400, 300, 180, 1, "全部", "是", 0.5, "专业游泳教学", "上架"],
            ["C005", "动感单车团课", "动感单车", "团课", 45, 0.3, 80, 60, 30, 30, "全部", "是", 0.3, "燃脂动感单车", "上架"],
            ["C006", "拳击私教课", "拳击", "1对1私教", 60, 1, 350, 250, 180, 1, "全部", "是", 0.5, "拳击训练", "上架"],
            ["C007", "拉伸康复课", "拉伸康复", "1对1私教", 60, 1, 280, 200, 180, 1, "全部", "否", 0, "运动后拉伸康复", "上架"],
            ["C008", "有氧操大班课", "有氧操", "大班课(9人+)", 60, 0.5, 100, 80, 90, 20, "全部", "是", 0.5, "燃脂有氧操", "上架"],
            ["C009", "体态矫正课", "体态矫正", "1对1私教", 60, 1, 320, 250, 180, 1, "全部", "否", 0, "专业体态评估与矫正", "上架"],
            ["C010", "体验课", "力量训练", "体验课", 30, 0.5, 99, 99, 7, 1, "全部", "是", 0.5, "新人体验课程", "上架"],
        ]
        for i, c in enumerate(courses):
            row = DATA_START_ROW + i
            for j, v in enumerate(c):
                ws.cell(row=row, column=j + 1, value=v)

    def _add_sample_staff(self):
        """添加示例员工"""
        ws = self.engine.get_sheet(SHEETS["staff"])
        today = date.today()
        staff = [
            ["E001", "张教练", "男", "健身教练", date(2020, 3, 1), date(1992, 5, 15), None, "13800138001", "zhang@fit.com", "在职", None,
             "NSCA-CPT, 康复认证", "私教课,拉伸康复",
             0.08, 0.07, 4000, 0, 0, 0, 0, 0, 0, 0, 0],
            ["E002", "李教练", "男", "瑜伽教练", date(2021, 6, 1), date(1994, 8, 20), None, "13800138002", "li@fit.com", "在职", None,
             "RYT200, 普拉提认证", "瑜伽,普拉提",
             0.07, 0.07, 3500, 0, 0, 0, 0, 0, 0, 0, 0],
            ["E003", "王销售", "女", "销售顾问", date(2022, 1, 15), date(1996, 3, 10), None, "13800138003", "wang@fit.com", "在职", None,
             "销售培训证书", "",
             0.10, 0, 3000, 0, 0, 0, 0, 0, 0, 0, 0],
            ["E004", "赵教练", "女", "游泳教练", date(2021, 9, 1), date(1993, 11, 5), None, "13800138004", "zhao@fit.com", "在职", None,
             "游泳救生证, 游泳教练证", "游泳",
             0.08, 0.07, 3800, 0, 0, 0, 0, 0, 0, 0, 0],
            ["E005", "刘教练", "男", "操课教练", date(2022, 3, 1), date(1995, 7, 22), None, "13800138005", "liu@fit.com", "在职", None,
             "有氧操认证, 单车认证", "动感单车,有氧操",
             0.06, 0.06, 3200, 0, 0, 0, 0, 0, 0, 0, 0],
            ["E006", "陈店长", "女", "店长", date(2019, 5, 1), date(1990, 2, 14), None, "13800138006", "chen@fit.com", "在职", None,
             "运营管理证书", "",
             0.05, 0, 6000, 0, 0, 0, 0, 0, 0, 0, 0],
        ]
        for i, s in enumerate(staff):
            row = DATA_START_ROW + i
            for j, v in enumerate(s):
                ws.cell(row=row, column=j + 1, value=v)

    def _add_sample_members(self):
        """添加示例会员"""
        ws = self.engine.get_sheet(SHEETS["member"])
        today = date.today()
        members = [
            ["M20260430001", "张三", "男", "13900139001", date(1990, 3, 15), None, "zhangsan@mail.com",
             "金卡", date(2025, 10, 1), date(2026, 10, 1), "有效",
             8800, 30, 5, 25, 0, 25, date(2026, 4, 28), 2, 10, "私教课", "李四", "", "正常", None, ""],
            ["M20260430002", "李女士", "女", "13900139002", date(1995, 7, 22), None, "lili@mail.com",
             "银卡", date(2026, 1, 15), date(2026, 7, 15), "有效",
             4500, 20, 3, 17, 0, 17, date(2026, 4, 25), 5, 8, "瑜伽", "", "", "正常", None, ""],
            ["M20260430003", "王五", "男", "13900139003", date(1988, 11, 8), None, "wangwu@mail.com",
             "普通", date(2026, 3, 1), date(2026, 9, 1), "有效",
             3000, 12, 0, 12, 0, 12, None, None, 0, "", "", "", "正常", None, ""],
            ["M20260430004", "赵六", "男", "13900139004", date(2000, 1, 1), None, "zhaoliu@mail.com",
             "钻石", date(2024, 6, 1), date(2027, 6, 1), "有效",
             20000, 60, 20, 40, 5, 35, date(2026, 4, 20), 10, 20, "私教课", "", "高水平会员", "正常", None, ""],
            ["M20260430005", "孙七", "女", "13900139005", date(1992, 5, 20), None, "sunqi@mail.com",
             "银卡", date(2025, 8, 1), date(2026, 5, 15), "即将到期",
             3600, 15, 12, 3, 0, 3, date(2026, 3, 15), 46, 12, "瑜伽", "", "课时即将用完", "需回访", None, ""],
            ["M20260430006", "周八", "女", "13900139006", date(1985, 12, 10), None, "zhouba@mail.com",
             "金卡", date(2025, 1, 1), date(2026, 1, 1), "已到期",
             12000, 48, 45, 3, 0, 3, date(2025, 12, 20), 130, 45, "普拉提", "", "长期未续费", "流失预警", None, ""],
        ]
        for i, m in enumerate(members):
            row = DATA_START_ROW + i
            for j, v in enumerate(m):
                ws.cell(row=row, column=j + 1, value=v)

    def _add_sample_sales(self):
        """添加示例售课记录"""
        ws = self.engine.get_sheet(SHEETS["sale"])
        today = date.today()
        sales = [
            ["S202604010001", date(2026, 4, 1), "王销售", "M20260430001", "张三", "金卡",
             "C001", "1对1私教课", "力量训练", "1对1私教",
             10, 0, 10, 300, 0.9, 2700, 2700, "微信", 0.1, 270, 180, date(2026, 9, 28),
             25, "是", "INV-20260401", "CT-20260401", "老客续费", "春季活动", ""],
            ["S202604010002", date(2026, 4, 1), "王销售", "M20260430002", "李女士", "银卡",
             "C002", "瑜伽小班课", "瑜伽", "小班课(3-8人)",
             20, 2, 22, 150, 0.95, 2850, 2800, "支付宝", 0.1, 280, 90, date(2026, 6, 30),
             17, "是", "INV-20260402", "CT-20260402", "新客到店", "新客优惠", ""],
            ["S202604150003", date(2026, 4, 15), "王销售", "M20260430003", "王五", "普通",
             "C010", "体验课", "力量训练", "体验课",
             1, 0, 1, 99, 1, 99, 99, "微信", 0.1, 9.9, 7, date(2026, 4, 22),
             12, "否", "", "", "线上推广", "美团体验", ""],
            ["S202603010004", date(2026, 3, 1), "张教练", "M20260430004", "赵六", "钻石",
             "C003", "普拉提器械课", "普拉提", "1对1私教",
             20, 5, 25, 350, 0.85, 5950, 5800, "银行卡", 0.08, 464, 180, date(2026, 8, 28),
             40, "是", "INV-20260301", "CT-20260301", "转介绍", "", ""],
            ["S202603150005", date(2026, 3, 15), "王销售", "M20260430005", "孙七", "银卡",
             "C002", "瑜伽小班课", "瑜伽", "小班课(3-8人)",
             15, 0, 15, 150, 0.9, 2025, 2000, "储值卡扣款", 0.1, 200, 90, date(2026, 6, 13),
             3, "是", "INV-20260315", "CT-20260315", "老客续费", "", ""],
        ]
        for i, s in enumerate(sales):
            row = DATA_START_ROW + i
            for j, v in enumerate(s):
                ws.cell(row=row, column=j + 1, value=v)

    def _add_sample_class_records(self):
        """添加示例上课记录"""
        ws = self.engine.get_sheet(SHEETS["class_record"])
        records = [
            ["CL202604010001", date(2026, 4, 1), "09:00", "10:00", "张教练", "", "M20260430001", "张三",
             "C001", "1对1私教课", "力量训练", "1对1私教",
             1, "已完成", date(2026, 3, 30), "", "非常满意", "很好", "体重75kg,体脂18%", "胸肌训练",
             "背部训练", 0.07, 70, 0, "S202604010001", ""],
            ["CL202604010002", date(2026, 4, 1), "10:30", "11:30", "李教练", "", "M20260430002", "李女士",
             "C002", "瑜伽小班课", "瑜伽", "小班课(3-8人)",
             1, "已完成", date(2026, 3, 31), "", "满意", "柔韧性有提升", "", "流瑜伽序列",
             "阴瑜伽", 0.07, 70, 0, "S202604010002", ""],
            ["CL202604020003", date(2026, 4, 2), "09:00", "10:00", "张教练", "", "M20260430001", "张三",
             "C001", "1对1私教课", "力量训练", "1对1私教",
             1, "已完成", date(2026, 4, 1), "", "非常满意", "状态不错", "", "背部训练",
             "腿部训练", 0.07, 70, 0, "S202604010001", ""],
            ["CL202604030004", date(2026, 4, 3), "08:00", "09:00", "李教练", "", "M20260430002", "李女士",
             "C002", "瑜伽小班课", "瑜伽", "小班课(3-8人)",
             1, "已完成", date(2026, 4, 2), "", "满意", "", "", "阴瑜伽",
             "流瑜伽", 0.07, 70, 0, "S202604010002", ""],
            ["CL202604200005", date(2026, 4, 20), "14:00", "15:00", "赵教练", "", "M20260430004", "赵六",
             "C003", "普拉提器械课", "普拉提", "1对1私教",
             1, "已完成", date(2026, 4, 19), "", "非常满意", "核心力量增强", "体重72kg", "核心训练",
             "全身整合训练", 0.07, 70, 0, "S202603010004", ""],
            ["CL202604250006", date(2026, 4, 25), "10:00", "11:00", "李教练", "", "M20260430002", "李女士",
             "C002", "瑜伽小班课", "瑜伽", "小班课(3-8人)",
             1, "已完成", date(2026, 4, 24), "", "满意", "进步明显", "", "流瑜伽进阶",
             "冥想放松", 0.07, 70, 0, "S202604010002", ""],
            ["CL202604280007", date(2026, 4, 28), "09:00", "10:00", "张教练", "", "M20260430001", "张三",
             "C001", "1对1私教课", "力量训练", "1对1私教",
             1, "已完成", date(2026, 4, 27), "", "非常满意", "训练效果显著", "", "腿部训练",
             "肩部训练", 0.07, 70, 0, "S202604010001", ""],
        ]
        for i, r in enumerate(records):
            row = DATA_START_ROW + i
            for j, v in enumerate(r):
                ws.cell(row=row, column=j + 1, value=v)

    # ==================== 内部工具方法 ====================

    def _write_title(self, ws, title):
        """写入Sheet标题"""
        ws.cell(row=1, column=1, value=f"{VERSION_TAG} - {title}")
        ws.cell(row=1, column=1).font = self._title_font()
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    def _write_headers(self, ws, sheet_key):
        """写入表头"""
        headers = HEADERS.get(sheet_key, [])
        for i, h in enumerate(headers):
            ws.cell(row=DATA_START_ROW - 1, column=i + 1, value=h)

    def _add_dropdowns(self, ws, sheet_key, col_options):
        """添加下拉列表"""
        headers = HEADERS.get(sheet_key, [])
        max_col = len(headers)
        max_row = 1000  # 留出足够行数

        for col, options in col_options.items():
            if col > max_col:
                continue
            col_letter = chr(64 + col) if col <= 26 else "A"
            cell_range = f"{col_letter}{DATA_START_ROW}:{col_letter}{max_row}"
            # 注意：sheet_key是内部键名，需要转为对应的中文Sheet名
            actual_sheet_name = ws.title
            self.engine.add_dropdown(actual_sheet_name, cell_range, options)

    def _set_col_widths(self, ws, sheet_key, widths):
        """设置列宽"""
        for col, width in widths.items():
            col_letter = chr(64 + col) if col <= 26 else "A"
            ws.column_dimensions[col_letter].width = width

    def _title_font(self):
        from openpyxl.styles import Font
        return Font(name="微软雅黑", bold=True, size=14, color="1F4E79")

    def _subtitle_font(self):
        from openpyxl.styles import Font
        return Font(name="微软雅黑", bold=True, size=12, color="2E75B6")

    def _label_font(self):
        from openpyxl.styles import Font
        return Font(name="微软雅黑", bold=True, size=10, color="333333")

    def _link_font(self):
        from openpyxl.styles import Font
        return Font(name="微软雅黑", bold=True, size=10, color="0563C1", underline="single")

    def _header_font(self):
        from openpyxl.styles import Font
        return Font(name="微软雅黑", bold=True, size=10, color="FFFFFF")


def main():
    """主函数"""
    # 确保data目录存在
    os.makedirs(DATA_DIR, exist_ok=True)

    # 初始化
    engine = ExcelEngine(EXCEL_PATH)
    initializer = WorkbookInitializer(engine)
    initializer.initialize(with_sample_data=True)
    engine.close()

    print("\n🎉 工作簿初始化成功！")
    print(f"📁 路径: {EXCEL_PATH}")
    print("💡 运行 python main.py 启动系统")


if __name__ == "__main__":
    main()
